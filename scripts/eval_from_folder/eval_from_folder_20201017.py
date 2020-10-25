import numpy as np
from PIL import Image
import sys
from eval.evaler import eval_memo
import json
import os
from random import shuffle
import shutil
import datetime
from kid.models.inception import InceptionV3
from kid.kid_score import calculate_kid_given_paths
from kid.fid_score import calculate_fid_given_paths
from skimage_scores import scores_from_floder as mse_etc_scores_from_floder
import openpyxl

from eval.utils import make_dataset,get_inner_path

''''''

class ret_class:
    def __init__(self):
        self.datetime = None
        self.floder_true = None
        self.floder_fake = None
        self.kid_mean=None
        self.fid=None
        self.kNN=None
        self.K_MMD=None
        self.WD=None
        # scores not use official
        self._IS=None
        self._MS=None
        self._mse_skimage=None
        self._ssim_skimage=None
        self._ssimrgb_skimage = None
        self._psnr_skimage=None
        self._kid_std = None
        self._fid_inkid_mean=None
        self._fid_inkid_std=None

if __name__=='__main__':
    if len(sys.argv) > 1:
        real_paths = sys.argv[1]
        fake_paths = sys.argv[2]
    else:
        real_paths = [r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.1.0_seed0_p2pHDmodel/real_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.1.1_seed0_p2pHDmodel_200epoch/real_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.0.0_seed0_jointDL1_connect_featuremap/real_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.3.0_seed0_dlv3p_moreparms/real_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.4.0_seed0_p2pHD_moreparms/real_result",
                      ]
        fake_paths = [r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.1.0_seed0_p2pHDmodel/fake_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.1.1_seed0_p2pHDmodel_200epoch/fake_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.0.0_seed0_jointDL1_connect_featuremap/fake_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.3.0_seed0_dlv3p_moreparms/fake_result",
                      r"/home/liangshuaizhe/code/GAN/programs_and_shfiles/AIDOMG_all_log_14.0.4.0_seed0_p2pHD_moreparms/fake_result",
                      ]

    score_filename = f'result_eval_{datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")}.json'
    score_xlsxname = f'result_eval_{datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")}.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet('Sheet')
    sh = wb['Sheet']
    # with open(os.path.join(os.getcwd(), score_filename), 'a') as f:
    #     json.dump(str(datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")), f)
    #     f.write('\n')
    #     json.dump(real_path, f)
    #     f.write('\n')
    #     json.dump(fake_path, f)
    #     f.write('\n')

    if not isinstance(real_paths,list):
        real_paths=[real_paths]
        fake_paths=[fake_paths]

    assert len(real_paths)==len(fake_paths)
    rets=[]
    for i in range(len(real_paths)):
        print(f"folder No.{i+1} processing...")
        real_path=real_paths[i]
        fake_path=fake_paths[i]

        imgsA = make_dataset(real_path)
        imgsB = make_dataset(fake_path)
        rets.append(ret_class())

        rets[i].datetime=datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        rets[i].floder_true=real_path
        rets[i].floder_fake = fake_path

        print(f'计算第{i+1}组的GAN六法')
        for s in [ 'inception_v3', 'resnet18']:#[ 'inception_v3','vgg13' ,'vgg16', 'vgg19', 'resnet18', 'resnet34', 'resnet50', 'resnet101', 'resnet152']:
            evaler=eval_memo(len(imgsA),conv_models=[s],gpu='0',needinception=True,needmode=True,needwasserstein=True) # -1 is cpu
            for j, f in enumerate(imgsA):
                img = np.array(Image.open(f))
                img=np.expand_dims(img,axis=0)
                evaler.add_imgA(img)
                print("\r%d/%d" % (j, len(imgsA)),end=" ")
            print("\n")
            for j, f in enumerate(imgsB):
                img = np.array(Image.open(f))
                img = np.expand_dims(img, axis=0)
                evaler.add_imgB(img)
                print("\r%d/%d" % (j, len(imgsB)),end=" ")
            print("\n")

            score=evaler.get_score()
            # print(score)
            if s=='inception_v3':
                rets[i]._IS=score['inception_v3']['inception']
                rets[i]._MS=score['inception_v3']['mode']
            elif s=='resnet18':
                rets[i].kNN=score["resnet18"]["knn"]["conv"]
                rets[i].K_MMD = score["resnet18"]["mmd"]["conv"]
                rets[i].WD=score["resnet18"]["wasserstein"]["conv"]

            # with open(os.path.join(os.getcwd(), score_filename), 'a') as f:
            #     json.dump(score, f)
            #     f.write('\n')


        print(f'计算第{i+1}组的fid')
        from fid.fid_score import fid_score
        fid = fid_score(real_path=real_path, fake_path=fake_path, gpu='0') # '' is cpu
        rets[i].fid=fid

        print(f'计算第{i+1}组的kid')
        paths_kid = [real_path] + [fake_path]
        results_kid = calculate_kid_given_paths(paths_kid, batch_size=50, cuda=True, dims=2048, model_type='inception')
        rets[i].kid_mean=results_kid[0][1]
        rets[i]._kid_std = results_kid[0][2]
        results_fid_inkid = calculate_fid_given_paths(paths_kid, batch_size=50, cuda=True, dims=2048, model_type='inception')
        rets[i]._fid_inkid_mean = results_fid_inkid[0][1]
        rets[i]._fid_inkid_std = results_fid_inkid[0][2]

        print(f'计算第{i + 1}组的mse等')
        ssim, mse, psnr,ssim_rgb = mse_etc_scores_from_floder(real_path, fake_path)
        rets[i]._ssim_skimage,rets[i]._mse_skimage,rets[i]._psnr_skimage,rets[i]._ssimrgb_skimage=ssim, mse, psnr,ssim_rgb

        with open(score_filename, 'a') as f:
            json.dump(rets[i].__dict__, f)
            f.write('\n')

        # 绘制表格
        titles=['datetime','floder_true','floder_fake','kid_mean','fid','kNN','K_MMD','WD','','_IS','_MS','_mse_skimage',
                '_ssim_skimage','_ssimrgb_skimage','_psnr_skimage','_kid_std','_fid_inkid_mean','_fid_inkid_std']

        row=i+2
        for col in range(1,len(titles)+1):
            sh.cell(row=row, column=col, value='' if titles[col-1]=='' else eval('rets[i].'+titles[col-1]))
        wb.save(score_xlsxname)